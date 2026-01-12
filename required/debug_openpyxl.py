
import sys
import os
from pathlib import Path

# Add project root to path
sys.path.append(str(Path(__file__).parent.parent))

try:
    from openpyxl import load_workbook
    print("openpyxl is installed.")
except ImportError:
    print("openpyxl is NOT installed.")
    sys.exit(1)

file_path = Path(r"c:\Users\asus\OneDrive\Documents\prmes_w\prmes_ww\required\templates\Class_data3.xlsx")

if not file_path.exists():
    print(f"File not found: {file_path}")
    sys.exit(1)

try:
    wb = load_workbook(file_path, data_only=True)
    print("Successfully loaded workbook.")
    print(f"Sheet names: {wb.sheetnames}")
except Exception as e:
    print(f"Failed to load workbook: {e}")
