#!/usr/bin/env python3
"""
Custom import script for 2018 Project Data
Imports both Review 1 and Review 2 from a consolidated Excel file
"""

import os
import io
import openpyxl
from pathlib import Path
from app import create_app
from models import db, Student, Evaluation
from utils import normalize_header
from upload_helpers import map_excel_columns_to_criteria, get_criteria_key_map
from sqlalchemy import text

# Config
EXCEL_PATH = Path("required/Project_2018_Data_Ready.xlsx")

def import_sheet(app, ws, phase, review):
    print(f"\nImporting Phase {phase} Review {review} from sheet: {ws.title}")
    
    # Get headers
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), tuple())
    fieldnames = [str(v) if v is not None else "" for v in header_row]
    
    # Map columns
    key_map = {}
    for raw in fieldnames:
        h = normalize_header(raw)
        key_map[h] = raw
        if h in ("name", "student_name"): key_map["name"] = raw
        elif h in ("seat_no", "seatno", "usn"): key_map["seat_no"] = raw
        elif h in ("literature_survey", "literature"): key_map["literature_survey"] = raw
        elif h in ("problem", "problem_identification"): key_map["problem_identification"] = raw
        elif h in ("objectives",): key_map["objectives"] = raw
        elif h in ("methodology",): key_map["methodology"] = raw
        elif h in ("presentation", "project_presentation_skill"): key_map["presentation"] = raw
        elif h in ("qa", "question_and_answer_session"): key_map["question_and_answer_session"] = raw

    with app.app_context():
        # Clear existing for this phase/review
        db.session.execute(text(f"DELETE FROM evaluation WHERE phase = {phase} AND review_no = {review}"))
        db.session.commit()
        
        created = 0
        for row_vals in ws.iter_rows(min_row=2, values_only=True):
            row = {fieldnames[i]: row_vals[i] for i in range(len(fieldnames))}
            name = str(row.get(key_map["name"]) or "").strip()
            seat_no = str(row.get(key_map["seat_no"]) or "").strip()
            
            if not name or not seat_no: continue
            
            student = Student.query.filter_by(seat_no=seat_no).first()
            if not student:
                student = Student(name=name, seat_no=seat_no)
                db.session.add(student)
                db.session.flush()
            
            try:
                # Default to 0 for missing marks
                def clean_row(r, km):
                    clean = {}
                    for k, v in r.items():
                        # If value is from a columns we expect to be an int, handle None
                        clean[k] = v if v is not None else 0
                    return clean
                
                row_cleaned = clean_row(row, key_map)
                comp, m1, m2, g = map_excel_columns_to_criteria(row_cleaned, key_map, phase, review)
                total = sum(comp.values())
                
                ev = Evaluation(
                    phase=phase, review_no=review, total_marks=total,
                    criteria1=comp["criteria1"], criteria2=comp["criteria2"],
                    criteria3=comp["criteria3"], criteria4=comp["criteria4"],
                    member1_criteria1=m1["criteria1"], member1_criteria2=m1["criteria2"],
                    member1_criteria3=m1["criteria3"], member1_criteria4=m1["criteria4"],
                    member2_criteria1=m2["criteria1"], member2_criteria2=m2["criteria2"],
                    member2_criteria3=m2["criteria3"], member2_criteria4=m2["criteria4"],
                    guide_criteria1=g["criteria1"], guide_criteria2=g["criteria2"],
                    guide_criteria3=g["criteria3"], guide_criteria4=g["criteria4"],
                    student=student
                )
                db.session.add(ev)
                created += 1
            except Exception as e:
                print(f"Error for {name}: {e}")
                
        db.session.commit()
        print(f"Done. Imported {created} records.")

def main():
    if not EXCEL_PATH.exists():
        print(f"File not found: {EXCEL_PATH}")
        return
    
    app = create_app()
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    if "Review 1" in wb.sheetnames:
        import_sheet(app, wb["Review 1"], 1, 1)
    if "Review 2" in wb.sheetnames:
        import_sheet(app, wb["Review 2"], 1, 2)
    
    print("\nBatch Import Complete!")

if __name__ == "__main__":
    main()
