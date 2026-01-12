#!/usr/bin/env python3
"""
Script to re-upload Phase 2 Review 2 data from Excel file
This will delete existing P2R2 data and re-import with correct phase/review parameters
"""

import io
from pathlib import Path
from openpyxl import load_workbook
from app import create_app
from models import db, Student, Evaluation
from utils import normalize_header
from upload_helpers import map_excel_columns_to_criteria
from sqlalchemy import text

# Configuration
EXCEL_PATH = Path(r"C:\Users\asus\OneDrive\Documents\Class_data1.xlsx")
PHASE = 2
REVIEW = 2

def main():
    print("="*60)
    print(f"Re-uploading data for Phase {PHASE} Review {REVIEW}")
    print("="*60)
    
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found: {EXCEL_PATH}")
        return
    
    # Create app and load data
    app = create_app()
    
    with app.app_context():
        print(f"\nReading Excel file: {EXCEL_PATH}")
        
        try:
            wb = load_workbook(EXCEL_PATH, data_only=True)
            ws = wb.active
            
            # Get headers
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), tuple())
            fieldnames = [str(v) if v is not None else "" for v in header_row]
            headers = [normalize_header(h) for h in fieldnames]
            
            print(f"Headers found: {fieldnames}")
            
            # Map columns
            key_map = {}
            for raw in fieldnames:
                h = normalize_header(raw)
                if h in ("name", "student_name"):
                    key_map["name"] = raw
                elif h in ("seat_no", "seatno", "usn", "univ_seat_no", "seat_no"):
                    key_map["seat_no"] = raw
                elif h in ("group", "group_no", "group_number", "project_group"):
                    key_map["group_no"] = raw
                elif h in ("project_title", "title"):
                    key_map["project_title"] = raw
                elif h in ("project_guide", "guide"):
                    key_map["project_guide"] = raw
                elif h in ("member1", "member_1", "chairperson"):
                    key_map["member1"] = raw
                elif h in ("member2", "member_2"):
                    key_map["member2"] = raw
                elif h in ("internal_guide", "guide", "internal_guide"):
                    key_map["internal_guide"] = raw
            
            print(f"\nMapped columns: {key_map}")
            
            missing = [k for k in ("name", "seat_no") if k not in key_map]
            if missing:
                print(f"ERROR: Missing required columns: {missing}")
                return
            
            # Check if we have three evaluators
            if not all(k in key_map for k in ("member1", "member2", "internal_guide")):
                print("ERROR: Excel must have Member 1, Member 2, and Internal Guide columns")
                return
            
            # Delete existing Phase 2 Review 2 data
            print(f"\nDeleting existing Phase {PHASE} Review {REVIEW} data...")
            count = db.session.execute(
                text(f"SELECT COUNT(*) FROM evaluation WHERE phase = {PHASE} AND review_no = {REVIEW}")
            ).scalar()
            print(f"Found {count} existing records")
            
            db.session.execute(
                text(f"DELETE FROM evaluation WHERE phase = {PHASE} AND review_no = {REVIEW}")
            )
            db.session.commit()
            print("Deleted existing records")
            
            # Import new data
            print(f"\nImporting new data for Phase {PHASE} Review {REVIEW}...")
            created = 0
            
            for row_vals in ws.iter_rows(min_row=2, values_only=True):
                row = {}
                for i, key in enumerate(fieldnames):
                    val = row_vals[i] if i < len(row_vals) else None
                    row[key] = val
                
                name = str(row.get(key_map["name"]) or "").strip()
                seat_no = str(row.get(key_map["seat_no"]) or "").strip()
                
                if not name or not seat_no:
                    continue
                
                group_no = str(row.get(key_map.get("group_no", ""), "") or "").strip()
                project_title = str(row.get(key_map.get("project_title", ""), "") or "").strip()
                project_guide = str(row.get(key_map.get("project_guide", ""), "") or "").strip()
                
                # Get or create student
                student = Student.query.filter_by(seat_no=seat_no).one_or_none()
                if not student:
                    student = Student(
                        name=name,
                        seat_no=seat_no,
                        group_no=group_no,
                        project_title=project_title,
                        project_guide=project_guide
                    )
                    db.session.add(student)
                else:
                    if group_no:
                        student.group_no = group_no
                    if project_title:
                        student.project_title = project_title
                    if project_guide:
                        student.project_guide = project_guide
                
                # Map columns to criteria using correct phase/review
                try:
                    comp, member1_comp, member2_comp, guide_comp = map_excel_columns_to_criteria(
                        row, key_map, PHASE, REVIEW
                    )
                    total = sum(comp.values())
                    
                    print(f"  {name} ({seat_no}): {list(comp.values())} = {total}")
                    
                except Exception as e:
                    print(f"ERROR processing {name}: {e}")
                    continue
                
                # Create new evaluation
                evaluation = Evaluation(
                    phase=PHASE,
                    review_no=REVIEW,
                    total_marks=total,
                    criteria1=comp["criteria1"],
                    criteria2=comp["criteria2"],
                    criteria3=comp["criteria3"],
                    criteria4=comp["criteria4"],
                    member1_criteria1=member1_comp["criteria1"],
                    member1_criteria2=member1_comp["criteria2"],
                    member1_criteria3=member1_comp["criteria3"],
                    member1_criteria4=member1_comp["criteria4"],
                    member2_criteria1=member2_comp["criteria1"],
                    member2_criteria2=member2_comp["criteria2"],
                    member2_criteria3=member2_comp["criteria3"],
                    member2_criteria4=member2_comp["criteria4"],
                    guide_criteria1=guide_comp["criteria1"],
                    guide_criteria2=guide_comp["criteria2"],
                    guide_criteria3=guide_comp["criteria3"],
                    guide_criteria4=guide_comp["criteria4"],
                    student=student,
                )
                db.session.add(evaluation)
                created += 1
            
            db.session.commit()
            
            print(f"\n{'='*60}")
            print(f"✓ Successfully imported {created} records for Phase {PHASE} Review {REVIEW}")
            print(f"{'='*60}")
            
            # Verify first record
            first_ev = Evaluation.query.filter_by(phase=PHASE, review_no=REVIEW).first()
            if first_ev:
                print(f"\nVerification - First record:")
                print(f"  Criteria: {first_ev.criteria1}, {first_ev.criteria2}, {first_ev.criteria3}, {first_ev.criteria4}")
                print(f"  Total: {first_ev.total_marks}")
                print(f"  Expected for P2R2: 9, 9, 14, 14 = 46")
                if [first_ev.criteria1, first_ev.criteria2, first_ev.criteria3, first_ev.criteria4] == [9, 9, 14, 14]:
                    print(f"  ✓ CORRECT!")
                else:
                    print(f"  ✗ Still wrong!")
            
        except PermissionError:
            print("\nERROR: Excel file is open. Please close it and try again.")
        except Exception as e:
            print(f"\nERROR: {type(e).__name__}: {e}")
            import traceback
            traceback.print_exc()

if __name__ == "__main__":
    main()
