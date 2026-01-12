import openpyxl
from pathlib import Path

def transform():
    src_path = Path("required/project_2018_data.xlsx")
    if not src_path.exists():
        src_path = Path("project_2018_data.xlsx")
    
    if not src_path.exists():
        print("Source file not found!")
        return

    wb = openpyxl.load_workbook(src_path, data_only=True)
    out_wb = openpyxl.Workbook()
    
    # Process Review 1
    if "review 1" in wb.sheetnames:
        ws1 = wb["review 1"]
        out_ws1 = out_wb.active
        out_ws1.title = "Review 1"
        # Headers for P1R1
        out_ws1.append(["Name", "Seat_no", "Literature Survey", "Problem Identification", "Project presentation skill", "Question and answer session"])
        
        # Data starts at row 11 in original
        for row_idx in range(11, ws1.max_row + 1):
            row = [cell.value for cell in ws1[row_idx]]
            if len(row) >= 7 and row[1]: # USN must exist
                out_ws1.append([row[2], row[1], row[3], row[4], row[5], row[6]])
        print("Added Review 1 sheet")

    # Process Review 2
    if "review 2" in wb.sheetnames:
        ws2 = wb["review 2"]
        out_ws2 = out_wb.create_sheet("Review 2")
        # Headers for P1R2
        out_ws2.append(["Name", "Seat_no", "Objectives", "Methodology", "Project presentation skill", "Question and answer session"])
        
        # Data starts at row 11 in original
        for row_idx in range(11, ws2.max_row + 1):
            row = [cell.value for cell in ws2[row_idx]]
            if len(row) >= 7 and row[1]: # USN must exist
                out_ws2.append([row[2], row[1], row[3], row[4], row[5], row[6]])
        print("Added Review 2 sheet")

    out_wb.save("required/Project_2018_Data_Ready.xlsx")
    print("Generated required/Project_2018_Data_Ready.xlsx with both reviews.")

if __name__ == "__main__":
    transform()
