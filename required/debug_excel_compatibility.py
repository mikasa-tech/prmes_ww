import sys
import os
from openpyxl import load_workbook

# Add current directory to path to import utils logic if needed, 
# but for isolation we'll just copy the relevant logic from app.py to be sure we test exactly what runs.

def normalize_header(h):
    if not h:
        return ""
    return str(h).lower().strip().replace(" ", "_").replace(".", "").replace("/", "_").replace("(", "").replace(")", "")

def check_file(filepath):
    print(f"Checking file: {filepath}")
    if not os.path.exists(filepath):
        print("  [ERROR] File not found.")
        return

    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), tuple())
        fieldnames = [str(v) if v is not None else "" for v in header_row]
        print(f"  Raw Headers: {fieldnames}")
        
        headers = [normalize_header(h) for h in fieldnames]
        # print(f"  Norm Headers: {headers}")

        key_map = {}
        for raw in fieldnames or []:
            h = normalize_header(raw)
            if h in ("name", "student_name"):
                key_map["name"] = raw
            elif h in ("seat_no", "seatno", "usn", "univ_seat_no"):
                key_map["seat_no"] = raw
            elif h in ("total", "total_marks", "average", "avg"):
                key_map["total"] = raw
            elif h in ("group", "group_no", "group_number", "project_group"):
                key_map["group_no"] = raw
            elif h in ("project_title", "title"):
                key_map["project_title"] = raw
            elif h in ("project_guide",):
                key_map["project_guide"] = raw
            elif h in ("member1", "member_1", "chairperson", "member1_total"):
                key_map["member1"] = raw
            elif h in ("member2", "member_2", "member2_total"):
                key_map["member2"] = raw
            elif h in ("internal_guide", "guide", "project_guide"): # Note: duplicate check in app.py logic?
                key_map["internal_guide"] = raw
            elif h in ("literature", "literature_survey"):
                key_map["literature_survey"] = raw
            elif h in ("problem", "problem_identification"):
                key_map["problem_identification"] = raw
            elif h in ("presentation", "project_presentation_skill", "presentation_skill"):
                key_map["presentation"] = raw
            elif h in ("qa", "qna", "question_answer", "question_and_answer_session"):
                key_map["question_answer"] = raw

        missing = [k for k in ("name", "seat_no") if k not in key_map]
        
        has_components = all(k in key_map for k in ("literature_survey","problem_identification","presentation","question_answer"))
        has_three_evaluators = all(k in key_map for k in ("member1","member2","internal_guide"))
        
        if "total" not in key_map and not has_components and not has_three_evaluators:
            missing.append("total or all four component columns or Member 1 + Member 2 + Internal Guide")
        
        if missing:
            print(f"  [FAIL] Missing columns: {', '.join(missing)}")
        else:
            print("  [PASS] File structure is valid.")
            if has_components:
                print("         - Detected component-wise structure.")
            if has_three_evaluators:
                print("         - Detected evaluator-wise structure.")
            if "total" in key_map:
                print("         - Detected total marks column.")

    except Exception as e:
        print(f"  [ERROR] Exception processing file: {e}")

if __name__ == "__main__":
    files = [
        "templates/Class_data3.xlsx",
        "../non_required/Class_data1.xlsx"
    ]
    for f in files:
        check_file(f)
