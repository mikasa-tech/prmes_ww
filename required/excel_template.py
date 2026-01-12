from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import date

def build_review1_workbook(student, ev):
    wb = Workbook()
    ws = wb.active
    ws.title = "Review-1"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Titles
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    c1 = ws.cell(row=1, column=1, value="FINAL YEAR STUDENTS MAJOR PROJECT WORK PHASE - 1")
    c1.font = Font(size=16, bold=True)
    c1.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
    c2 = ws.cell(row=2, column=1, value="CONTINUOUS INTERNAL EVALUATION (CIE) OF MAJOR PROJECT WORK PHASE - 1")
    c2.font = Font(size=13, bold=True)
    c2.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=12)
    c3 = ws.cell(row=3, column=1, value="REVIEW - 1")
    c3.font = Font(size=13, bold=True)
    c3.alignment = Alignment(horizontal="center")

    # Student header
    ws.merge_cells("A5:B5"); ws["A5"] = f"Group No: {student.group_no or '-'}"
    ws.merge_cells("C5:L5"); ws["C5"] = f"Project Title: {student.project_title or '-'}"

    # Table header
    ws.merge_cells("A7:A9"); ws["A7"] = "Sl. No"
    ws.merge_cells("B7:B9"); ws["B7"] = "Univ. Seat No."
    ws.merge_cells("C7:C9"); ws["C7"] = "Name of the Student"
    ws.merge_cells("D7:D9"); ws["D7"] = "Aspect for Assessment"
    ws.merge_cells("E7:H7"); ws["E7"] = "Continuous Internal Evaluation (CIE) by"
    ws.merge_cells("I7:I9"); ws["I7"] = "Average CIE Marks (50)"

    ws.merge_cells("E8:F8"); ws["E8"] = "Chairperson (Member-1)"
    ws["G8"] = "Member-2"
    ws["H8"] = "Internal Guide"

    ws["E9"] = "Marks"; ws["F9"] = ""; ws["G9"] = "Marks"; ws["H9"] = "Marks"

    # Rows for components
    start = 10
    components = [
        ("(a) Literature Survey (20 Marks)", ev.literature_survey, ev.member1_literature, ev.member2_literature, ev.guide_literature),
        ("(b) Problem Identification (10 Marks)", ev.problem_identification, ev.member1_problem, ev.member2_problem, ev.guide_problem),
        ("(c) Project presentation skill (10 Marks)", ev.presentation, ev.member1_presentation, ev.member2_presentation, ev.guide_presentation),
        ("(d) Question and answer session (10 Marks)", ev.question_answer, ev.member1_qa, ev.member2_qa, ev.guide_qa),
    ]

    for idx, (label, comp_mark, m1, m2, g) in enumerate(components):
        r = start + idx
        if idx == 0:
            ws.merge_cells(start_row=r, start_column=1, end_row=r+len(components)-1, end_column=1)
            ws.merge_cells(start_row=r, start_column=2, end_row=r+len(components)-1, end_column=2)
            ws.merge_cells(start_row=r, start_column=3, end_row=r+len(components)-1, end_column=3)
            ws.cell(row=r, column=1, value=1)
            ws.cell(row=r, column=2, value=student.seat_no)
            ws.cell(row=r, column=3, value=student.name)
        ws.cell(row=r, column=4, value=label)

        m1_val = m1 or comp_mark
        m2_val = m2 or comp_mark
        g_val = g or comp_mark

        ws.cell(row=r, column=5, value=m1_val)
        ws.cell(row=r, column=6, value="")
        ws.cell(row=r, column=7, value=m2_val)
        ws.cell(row=r, column=8, value=g_val)

        try:
            avg_val = int(round((int(m1_val or 0) + int(m2_val or 0) + int(g_val or 0)) / 3))
        except Exception:
            avg_val = None
        ws.cell(row=r, column=9, value=avg_val)

    total_row = start + len(components)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    ws.cell(row=total_row, column=1, value="Total Marks (50)")

    avg_total_val = 0
    for rr in range(start, start + len(components)):
        cell_val = ws.cell(row=rr, column=9).value or 0
        try:
            avg_total_val += int(cell_val)
        except Exception:
            pass
    ws.cell(row=total_row, column=9, value=avg_total_val)

    # Borders and alignment
    for row in ws.iter_rows(min_row=7, max_row=total_row, min_col=1, max_col=9):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border

    widths = [6, 14, 24, 40, 14, 6, 14, 14, 18]
    for i, w in enumerate(widths, start=1):
        col_letter = chr(ord('A') + i - 1)
        ws.column_dimensions[col_letter].width = w

    # Footer placeholders
    ws.cell(row=total_row+5, column=2, value="Project coordinator")
    ws.cell(row=total_row+5, column=5, value="Chairperson (Member - 1)")
    ws.cell(row=total_row+5, column=7, value="Member - 2")
    ws.cell(row=total_row+5, column=9, value="Guide")

    return wb


def build_summary_workbook(students_evaluations):
    """
    Build a summary Excel workbook for all students
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "CIE Review 1 Summary"
    
    # Define styles
    header_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    
    header_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    # Headers
    headers = [
        'Group', 'Seat No', 'Student Name', 'Project Title',
        'Literature\n(20)', 'Problem ID\n(10)', 'Presentation\n(10)', 'Q&A\n(10)', 'Total\n(50)'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col)].width = 12
    
    # Set specific column widths
    ws.column_dimensions['C'].width = 20  # Student Name
    ws.column_dimensions['D'].width = 25  # Project Title
    
    # Data rows
    row = 2
    for student, evaluation in students_evaluations:
        data = [
            student.group_no or "",
            student.seat_no,
            student.name,
            student.project_title or "",
            evaluation.literature_survey,
            evaluation.problem_identification,
            evaluation.presentation,
            evaluation.question_answer,
            evaluation.total_marks
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = normal_font
            cell.border = thin_border
            if col <= 4:  # Text columns
                cell.alignment = left_align
            else:  # Number columns
                cell.alignment = center_align
        row += 1
    
    return wb