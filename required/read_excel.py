#!/usr/bin/env python3

import openpyxl

def read_excel_guides():
    """Read guide information from Class_data1.xlsx"""
    try:
        wb = openpyxl.load_workbook('Class_data1.xlsx')
        ws = wb.active
        
        # Get header row
        headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        print("Headers:", headers)
        
        # Find the guide column
        guide_col_idx = None
        for i, header in enumerate(headers):
            if header and 'guide' in str(header).lower():
                guide_col_idx = i + 1  # Excel is 1-indexed
                print(f"Found guide column: '{header}' at position {guide_col_idx}")
                break
        
        if not guide_col_idx:
            print("No 'guide' column found!")
            return
        
        print("\nData with guides:")
        for row in range(2, ws.max_row + 1):
            row_data = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
            if any(cell for cell in row_data):  # Skip empty rows
                guide_name = ws.cell(row, guide_col_idx).value
                print(f"Row {row}: {row_data} | Guide: {guide_name}")
                
    except Exception as e:
        print(f"Error reading Excel: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    read_excel_guides()