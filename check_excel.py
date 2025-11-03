import openpyxl
from pathlib import Path

# Try to read the Excel file
excel_path = Path(r"C:\Users\asus\OneDrive\Documents\Class_data1.xlsx")

print(f"Checking file: {excel_path}")
print(f"File exists: {excel_path.exists()}")

try:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    print(f"\nSheets in workbook: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'='*60}")
        print(f"Sheet: {sheet_name}")
        print(f"{'='*60}")
        
        # Get headers
        headers = [cell.value for cell in ws[1]]
        print(f"Headers: {headers}")
        
        # Check if this looks like evaluation data
        if 'Member 1' in headers or 'Member1' in headers:
            print(f"\nThis sheet contains evaluation data!")
            print(f"Sample data (first student):")
            if ws.max_row >= 2:
                row_data = [cell.value for cell in ws[2]]
                print(f"Row 2: {row_data}")
                
                # Find Member 1, Member 2, Internal Guide columns
                member1_idx = headers.index('Member 1') if 'Member 1' in headers else (headers.index('Member1') if 'Member1' in headers else -1)
                member2_idx = headers.index('Member 2') if 'Member 2' in headers else (headers.index('Member2') if 'Member2' in headers else -1)
                guide_idx = headers.index('Internal Guide') if 'Internal Guide' in headers else -1
                
                if member1_idx >= 0 and member2_idx >= 0 and guide_idx >= 0:
                    print(f"\nEvaluator totals:")
                    print(f"  Member 1: {row_data[member1_idx]}")
                    print(f"  Member 2: {row_data[member2_idx]}")
                    print(f"  Internal Guide: {row_data[guide_idx]}")
        
        print(f"\nTotal rows: {ws.max_row}")
    
    wb.close()
    
except PermissionError:
    print("\n⚠️  ERROR: File is currently open in Excel or another program!")
    print("Please close the file and try again.")
except Exception as e:
    print(f"\n⚠️  ERROR: {type(e).__name__}: {e}")

print("\n" + "="*60)
print("IMPORTANT: Does your Excel file have:")
print("1. Only ONE sheet with data for ALL reviews?")
print("2. OR separate sheets for Phase1-R1, Phase1-R2, Phase2-R1, Phase2-R2?")
print("="*60)
