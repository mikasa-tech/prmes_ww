#!/usr/bin/env python3

import openpyxl
from models import db, Student, Evaluation
from app import create_app
from utils import reverse_engineer_components

def reimport_with_guides():
    """Re-import Excel file with project guide information"""
    app = create_app()
    
    with app.app_context():
        try:
            # Read Excel file
            wb = openpyxl.load_workbook('Class_data1.xlsx', data_only=True)
            ws = wb.active
            
            # Clear existing data
            db.session.execute(db.text("DELETE FROM evaluation"))
            db.session.execute(db.text("DELETE FROM student"))
            db.session.commit()
            print("Cleared existing data")
            
            # Get headers
            headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
            print("Headers:", headers)
            
            # Find column indices
            name_col = headers.index('Name') + 1
            seat_col = headers.index('Seat_no') + 1
            group_col = headers.index('Project Group') + 1
            title_col = headers.index('Project Title') + 1
            guide_col = headers.index('Project Guide') + 1
            member1_col = headers.index('Member 1') + 1
            member2_col = headers.index('Member 2') + 1
            internal_guide_col = headers.index('Internal Guide') + 1
            
            created_students = 0
            created_evaluations = 0
            
            # Process each row
            for row in range(2, ws.max_row + 1):
                name = str(ws.cell(row, name_col).value or "").strip()
                seat_no = str(ws.cell(row, seat_col).value or "").strip()
                group_no = str(ws.cell(row, group_col).value or "").strip()
                project_title = str(ws.cell(row, title_col).value or "").strip()
                project_guide = str(ws.cell(row, guide_col).value or "").strip()
                
                if not name or not seat_no:
                    continue
                
                print(f"Processing: {name} ({seat_no}) - Guide: {project_guide}")
                
                # Create student
                student = Student(
                    name=name,
                    seat_no=seat_no,
                    group_no=group_no,
                    project_title=project_title,
                    project_guide=project_guide
                )
                db.session.add(student)
                db.session.flush()  # Get student ID
                created_students += 1
                
                # Get evaluator marks
                member1_total = int(float(ws.cell(row, member1_col).value or 0))
                member2_total = int(float(ws.cell(row, member2_col).value or 0))
                guide_total = int(float(ws.cell(row, internal_guide_col).value or 0))
                
                # Calculate average
                total = int(round((member1_total + member2_total + guide_total) / 3))
                
                # Distribute components
                member1_comp = reverse_engineer_components(member1_total)
                member2_comp = reverse_engineer_components(member2_total)
                guide_comp = reverse_engineer_components(guide_total)
                
                # Calculate average components
                avg_lit = int(round((member1_comp["literature_survey"] + member2_comp["literature_survey"] + guide_comp["literature_survey"]) / 3))
                avg_prob = int(round((member1_comp["problem_identification"] + member2_comp["problem_identification"] + guide_comp["problem_identification"]) / 3))
                avg_pres = int(round((member1_comp["presentation"] + member2_comp["presentation"] + guide_comp["presentation"]) / 3))
                avg_qa = int(round((member1_comp["question_answer"] + member2_comp["question_answer"] + guide_comp["question_answer"]) / 3))
                
                # Create evaluation
                evaluation = Evaluation(
                    student_id=student.id,
                    review_no=1,
                    total_marks=avg_lit + avg_prob + avg_pres + avg_qa,
                    literature_survey=avg_lit,
                    problem_identification=avg_prob,
                    presentation=avg_pres,
                    question_answer=avg_qa,
                    # Individual evaluator marks
                    member1_literature=member1_comp["literature_survey"],
                    member1_problem=member1_comp["problem_identification"],
                    member1_presentation=member1_comp["presentation"],
                    member1_qa=member1_comp["question_answer"],
                    member2_literature=member2_comp["literature_survey"],
                    member2_problem=member2_comp["problem_identification"],
                    member2_presentation=member2_comp["presentation"],
                    member2_qa=member2_comp["question_answer"],
                    guide_literature=guide_comp["literature_survey"],
                    guide_problem=guide_comp["problem_identification"],
                    guide_presentation=guide_comp["presentation"],
                    guide_qa=guide_comp["question_answer"]
                )
                db.session.add(evaluation)
                created_evaluations += 1
            
            db.session.commit()
            print(f"\n✅ Successfully imported:")
            print(f"   - {created_students} students")
            print(f"   - {created_evaluations} evaluations")
            
        except Exception as e:
            db.session.rollback()
            print(f"❌ Error: {e}")
            import traceback
            traceback.print_exc()

if __name__ == "__main__":
    reimport_with_guides()