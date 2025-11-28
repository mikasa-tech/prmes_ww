import io
import os
import csv
from pathlib import Path
from datetime import date
from flask import Flask, render_template, request, redirect, url_for, send_file, flash
from models import db, Student, Evaluation
from utils import reverse_engineer_components, normalize_header, TOTAL_MAX
from pdf_template import build_review1_pdf
from comprehensive_pdf_template import build_comprehensive_pdf
from upload_helpers import map_excel_columns_to_criteria
from review_config import get_review_config
from dotenv import load_dotenv
import sqlalchemy
from sqlalchemy import text
from openpyxl import load_workbook

APP_DIR = Path(__file__).parent.resolve()
EXPORTS_DIR = APP_DIR / "exports"

def create_app() -> Flask:
    # Load .env if present
    load_dotenv()

    app = Flask(__name__)
    app.config["SECRET_KEY"] = "dev-secret"

    # Optional MySQL auto-creation if envs provided; fallback to SQLite
    db_user = os.getenv("DB_USER")
    db_pass = os.getenv("DB_PASS", "")
    db_host = os.getenv("DB_HOST")
    db_port = os.getenv("DB_PORT", "3306")
    db_name = os.getenv("DB_NAME")

    if db_host and db_user and db_name:
        # Ensure database exists
        engine_no_db = sqlalchemy.create_engine(
            f"mysql+pymysql://{db_user}:{db_pass}@{db_host}:{db_port}/?charset=utf8mb4",
            future=True,
        )
        with engine_no_db.connect() as conn:
            conn.execute(text(f"CREATE DATABASE IF NOT EXISTS `{db_name}` DEFAULT CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci;"))
        app.config["SQLALCHEMY_DATABASE_URI"] = (
            f"mysql+pymysql://{db_user}:{db_pass}@{db_host}:{db_port}/{db_name}?charset=utf8mb4"
        )
    else:
        app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{APP_DIR/'app.db'}"

    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
    db.init_app(app)
    with app.app_context():
        db.create_all()

    @app.route("/")
    def index():
        return redirect(url_for("upload_csv"))

    @app.route("/upload", methods=["GET", "POST"])
    def upload_csv():
        if request.method == "POST":
            # Get phase and review from form
            try:
                phase = int(request.form.get("phase", 1))
                review_no = int(request.form.get("review", 1))
            except (ValueError, TypeError):
                flash("Invalid phase or review number.", "error")
                return redirect(request.url)
            
            if phase not in [1, 2] or review_no not in [1, 2]:
                flash("Phase and Review must be 1 or 2.", "error")
                return redirect(request.url)
            
            file = request.files.get("file")
            if not file or file.filename.strip() == "":
                flash("Please choose an .xlsx file.", "error")
                return redirect(request.url)

            filename = file.filename.lower().strip()
            if not filename.endswith(".xlsx"):
                flash("Only .xlsx files are accepted.", "error")
                return redirect(request.url)

            # Read Excel (.xlsx) with openpyxl
            raw = file.stream.read()
            try:
                wb = load_workbook(io.BytesIO(raw), data_only=True)
            except Exception:
                flash("Could not read the .xlsx file. Please ensure it is a valid Excel file.", "error")
                return redirect(request.url)
            ws = wb.active

            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), tuple())
            fieldnames = [str(v) if v is not None else "" for v in header_row]
            headers = [normalize_header(h) for h in fieldnames]

            def iter_rows_dict():
                for vals in ws.iter_rows(min_row=2, values_only=True):
                    row = {}
                    for i, key in enumerate(fieldnames):
                        val = vals[i] if i < len(vals) else None
                        row[key] = val
                    yield row
            rows_iter = iter_rows_dict()

            # Map common header names to our keys
            # Required: student name, seat no. Optional: group no, project title, evaluator marks.
            key_map = {}
            for raw in fieldnames or []:
                h = normalize_header(raw)
                if h in ("name", "student_name"):
                    key_map["name"] = raw
                elif h in ("seat_no", "seatno", "usn", "univ_seat_no"):
                    key_map["seat_no"] = raw
                elif h in ("total", "total_marks", "average", "avg"):
                    key_map["total"] = raw
                elif h in ("group", "group_no", "group_number", "project_group"):
                    key_map["group_no"] = raw
                elif h in ("project_title", "title"):
                    key_map["project_title"] = raw
                elif h in ("project_guide",):
                    key_map["project_guide"] = raw
                elif h in ("member1", "member_1", "chairperson", "member1_total"):
                    key_map["member1"] = raw
                elif h in ("member2", "member_2", "member2_total"):
                    key_map["member2"] = raw
                elif h in ("internal_guide", "guide", "project_guide"):
                    key_map["internal_guide"] = raw
                elif h in ("literature", "literature_survey"):
                    key_map["literature_survey"] = raw
                elif h in ("problem", "problem_identification"):
                    key_map["problem_identification"] = raw
                elif h in ("presentation", "project_presentation_skill", "presentation_skill"):
                    key_map["presentation"] = raw
                elif h in ("qa", "qna", "question_answer", "question_and_answer_session"):
                    key_map["question_answer"] = raw

            missing = [k for k in ("name", "seat_no") if k not in key_map]
            has_components = all(k in key_map for k in ("literature_survey","problem_identification","presentation","question_answer"))
            has_three_evaluators = all(k in key_map for k in ("member1","member2","internal_guide"))
            if "total" not in key_map and not has_components and not has_three_evaluators:
                missing.append("total or all four component columns or Member 1 + Member 2 + Internal Guide")
            if missing:
                flash(f"Excel file missing required columns: {', '.join(missing)}", "error")
                return redirect(request.url)

            # Delete existing evaluations for this specific phase and review only
            try:
                existing_evals = Evaluation.query.filter_by(phase=phase, review_no=review_no).all()
                for ev in existing_evals:
                    db.session.delete(ev)
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                flash(f"Failed to delete existing Phase {phase} Review {review_no} data before import: {str(e)}", "error")
                return redirect(request.url)

            created = 0
            for row in rows_iter:
                name = str(row.get(key_map["name"]) or "").strip()
                seat_no = str(row.get(key_map["seat_no"]) or "").strip()
                if not name or not seat_no:
                    continue

                group_no = str(row.get(key_map.get("group_no", ""), "") or "").strip()
                project_title = str(row.get(key_map.get("project_title", ""), "") or "").strip()
                project_guide = str(row.get(key_map.get("project_guide", ""), "") or "").strip()

                student = Student.query.filter_by(seat_no=seat_no).one_or_none()
                if not student:
                    student = Student(name=name, seat_no=seat_no, group_no=group_no, project_title=project_title, project_guide=project_guide)
                    db.session.add(student)
                else:
                    # update all fields if provided (including name for corrections/updates)
                    if name:
                        student.name = name
                    if group_no:
                        student.group_no = group_no
                    if project_title:
                        student.project_title = project_title
                    if project_guide:
                        student.project_guide = project_guide

                # Use dynamic column mapping based on phase/review
                try:
                    comp, member1_comp, member2_comp, guide_comp = map_excel_columns_to_criteria(
                        row, key_map, phase, review_no
                    )
                    total = sum(comp.values())
                except Exception as e:
                    flash(f"Error processing row for {name}: {str(e)}", "error")
                    continue

                # Upsert: one evaluation per student per phase per review
                existing = Evaluation.query.filter_by(student=student, phase=phase, review_no=review_no).one_or_none()
                if existing:
                    existing.phase = phase
                    existing.total_marks = total
                    existing.criteria1 = comp["criteria1"]
                    existing.criteria2 = comp["criteria2"]
                    existing.criteria3 = comp["criteria3"]
                    existing.criteria4 = comp["criteria4"]
                    existing.member1_criteria1 = member1_comp["criteria1"]
                    existing.member1_criteria2 = member1_comp["criteria2"]
                    existing.member1_criteria3 = member1_comp["criteria3"]
                    existing.member1_criteria4 = member1_comp["criteria4"]
                    existing.member2_criteria1 = member2_comp["criteria1"]
                    existing.member2_criteria2 = member2_comp["criteria2"]
                    existing.member2_criteria3 = member2_comp["criteria3"]
                    existing.member2_criteria4 = member2_comp["criteria4"]
                    existing.guide_criteria1 = guide_comp["criteria1"]
                    existing.guide_criteria2 = guide_comp["criteria2"]
                    existing.guide_criteria3 = guide_comp["criteria3"]
                    existing.guide_criteria4 = guide_comp["criteria4"]
                else:
                    evaluation = Evaluation(
                        phase=phase,
                        review_no=review_no,
                        total_marks=total,
                        criteria1=comp["criteria1"],
                        criteria2=comp["criteria2"],
                        criteria3=comp["criteria3"],
                        criteria4=comp["criteria4"],
                        # Individual evaluator marks
                        member1_criteria1=member1_comp["criteria1"],
                        member1_criteria2=member1_comp["criteria2"],
                        member1_criteria3=member1_comp["criteria3"],
                        member1_criteria4=member1_comp["criteria4"],
                        member2_criteria1=member2_comp["criteria1"],
                        member2_criteria2=member2_comp["criteria2"],
                        member2_criteria3=member2_comp["criteria3"],
                        member2_criteria4=member2_comp["criteria4"],
                        guide_criteria1=guide_comp["criteria1"],
                        guide_criteria2=guide_comp["criteria2"],
                        guide_criteria3=guide_comp["criteria3"],
                        guide_criteria4=guide_comp["criteria4"],
                        student=student,
                    )
                    db.session.add(evaluation)
                    created += 1

            db.session.commit()

            # TODO: Update CSV export for new criteria-based system (temporarily disabled)

            if created > 0:
                flash(f"Imported {created} evaluation record(s) for Phase {phase} Review {review_no}.", "success")
            return redirect(url_for("list_students", phase=phase, review=review_no))

        return render_template("upload.html")

    @app.route("/students")
    def list_students():
        # Get phase and review from query params, default to Phase 1 Review 1
        phase = request.args.get("phase", 1, type=int)
        review = request.args.get("review", 1, type=int)
        
        # Get review configuration
        config = get_review_config(phase, review)
        
        students = Student.query.order_by(Student.group_no, Student.name).all()
        # Filter evaluations by phase and review
        data = []
        for s in students:
            ev = Evaluation.query.filter_by(student=s, phase=phase, review_no=review).first()
            if ev:
                data.append((s, ev))
        return render_template("students.html", items=data, phase=phase, review=review, config=config)
    
    @app.route("/students/groupwise")
    def students_groupwise():
        # Get phase and review from query params
        phase = request.args.get("phase", 1, type=int)
        review = request.args.get("review", 1, type=int)
        
        # Get review configuration
        config = get_review_config(phase, review)
        
        # Get all students with their evaluations, grouped by group_no
        students = Student.query.order_by(Student.group_no, Student.name).all()
        
        # Group students by group_no
        groups = {}
        for student in students:
            group_no = student.group_no or "No Group"
            if group_no not in groups:
                groups[group_no] = []
            
            # Get evaluation for specific phase and review
            ev = Evaluation.query.filter_by(student=student, phase=phase, review_no=review).first()
            if ev:
                groups[group_no].append((student, ev))
        
        return render_template("students_groupwise.html", groups=groups, phase=phase, review=review, config=config)
    
    @app.route("/students/guidewise")
    def students_guidewise():
        # Get phase and review from query params
        phase = request.args.get("phase", 1, type=int)
        review = request.args.get("review", 1, type=int)
        
        # Get review configuration
        config = get_review_config(phase, review)
        
        # Get all students with their evaluations, grouped by guide
        students = Student.query.order_by(Student.name).all()
        
        # Group students by their actual project guide from database
        guides = {}
        
        for student in students:
            ev = Evaluation.query.filter_by(student=student, phase=phase, review_no=review).first()
            if ev and student.project_guide:
                guide_name = student.project_guide.strip()
                
                if guide_name not in guides:
                    guides[guide_name] = []
                guides[guide_name].append((student, ev))
        
        return render_template("students_guidewise.html", guides=guides, phase=phase, review=review, config=config)
    
    @app.route("/students/individual")
    def students_individual():
        # Get phase and review from query params
        phase = request.args.get("phase", 1, type=int)
        review = request.args.get("review", 1, type=int)
        
        # Get review configuration
        config = get_review_config(phase, review)
        
        # Enhanced individual view with more details
        students = Student.query.order_by(Student.name).all()
        data = []
        for s in students:
            ev = Evaluation.query.filter_by(student=s, phase=phase, review_no=review).first()
            if ev:
                # Add more detailed individual data
                individual_data = {
                    'student': s,
                    'evaluation': ev,
                    'member1_total': (ev.member1_criteria1 or 0) + (ev.member1_criteria2 or 0) + 
                                   (ev.member1_criteria3 or 0) + (ev.member1_criteria4 or 0),
                    'member2_total': (ev.member2_criteria1 or 0) + (ev.member2_criteria2 or 0) + 
                                   (ev.member2_criteria3 or 0) + (ev.member2_criteria4 or 0),
                    'guide_total': (ev.guide_criteria1 or 0) + (ev.guide_criteria2 or 0) + 
                                  (ev.guide_criteria3 or 0) + (ev.guide_criteria4 or 0),
                }
                data.append(individual_data)
        
        return render_template("students_individual.html", students_data=data, phase=phase, review=review, config=config)

    @app.route("/students/<int:student_id>")
    def student_detail(student_id: int):
        # Get phase and review from query params
        phase = request.args.get("phase", 1, type=int)
        review = request.args.get("review", 1, type=int)
        
        student = Student.query.get_or_404(student_id)
        # Get evaluation for specific phase and review
        ev = Evaluation.query.filter_by(student=student, phase=phase, review_no=review).first()
        
        if not ev:
            flash(f"No evaluation found for Phase {phase} Review {review}", "error")
            return redirect(url_for("list_students", phase=phase, review=review))
        
        # Get review configuration
        config = get_review_config(phase, review)
        
        return render_template("student_detail.html", student=student, ev=ev, phase=phase, review=review, config=config)

    @app.route("/students/<int:student_id>/download")
    def download_review1(student_id: int):
        # Get phase and review from query params
        phase = request.args.get("phase", 1, type=int)
        review = request.args.get("review", 1, type=int)
        
        student = Student.query.get_or_404(student_id)
        # Get evaluation for specific phase and review
        ev = Evaluation.query.filter_by(student=student, phase=phase, review_no=review).first()
        
        if not ev:
            flash(f"No evaluation found for Phase {phase} Review {review}", "error")
            return redirect(url_for("list_students", phase=phase, review=review))
        
        pdf_buffer = build_review1_pdf(student, ev, phase, review)
        filename = f"Phase{phase}_Review{review}_{student.seat_no}_{student.name.replace(' ', '_')}.pdf"
        return send_file(pdf_buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")

    @app.route("/students/<int:student_id>/csv")
    def download_review1_csv(student_id: int):
        student = Student.query.get_or_404(student_id)
        ev = sorted(student.evaluations, key=lambda e: (e.review_no, e.id))[-1]

        # Prepare per-student CSV with component rows and averages
        bio = io.StringIO()
        writer = csv.writer(bio)
        writer.writerow(["Seat No", student.seat_no])
        writer.writerow(["Name", student.name])
        writer.writerow(["Group No", student.group_no or "-"])
        writer.writerow(["Project Title", student.project_title or "-"])
        writer.writerow([])
        writer.writerow(["Component", "Member 1", "Member 2", "Internal Guide", "Average"]) 

        rows = [
            ("Literature Survey (20)", ev.member1_literature or 0, ev.member2_literature or 0, ev.guide_literature or 0),
            ("Problem Identification (10)", ev.member1_problem or 0, ev.member2_problem or 0, ev.guide_problem or 0),
            ("Presentation (10)", ev.member1_presentation or 0, ev.member2_presentation or 0, ev.guide_presentation or 0),
            ("Q&A (10)", ev.member1_qa or 0, ev.member2_qa or 0, ev.guide_qa or 0),
        ]

        member1_total = 0
        member2_total = 0
        guide_total = 0
        avg_total = 0
        for label, m1, m2, g in rows:
            avg = int(round((int(m1) + int(m2) + int(g)) / 3))
            writer.writerow([label, m1, m2, g, avg])
            member1_total += int(m1)
            member2_total += int(m2)
            guide_total += int(g)
            avg_total += avg

        writer.writerow([])
        writer.writerow(["Total (50)", member1_total, member2_total, guide_total, avg_total])

        bio.seek(0)
        filename = f"Review-1_{student.seat_no}_{student.name.replace(' ', '_')}.csv"
        return send_file(io.BytesIO(bio.getvalue().encode("utf-8")), as_attachment=True, download_name=filename, mimetype="text/csv")

    @app.route("/export.csv")
    def download_export_csv():
        EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
        export_path = EXPORTS_DIR / "evaluations_review1.csv"
        if not export_path.exists():
            # If file missing (e.g., before any uploads), build it now from current DB
            with export_path.open("w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow([
                    "group_no",
                    "project_title",
                    "seat_no",
                    "name",
                    "review_no",
                    "member1_literature",
                    "member1_problem",
                    "member1_presentation",
                    "member1_qa",
                    "member1_total",
                    "member2_literature",
                    "member2_problem",
                    "member2_presentation",
                    "member2_qa",
                    "member2_total",
                    "guide_literature",
                    "guide_problem",
                    "guide_presentation",
                    "guide_qa",
                    "guide_total",
                    "avg_literature",
                    "avg_problem",
                    "avg_presentation",
                    "avg_qa",
                    "avg_total_marks",
                ])
                students = Student.query.order_by(Student.group_no, Student.name).all()
                for s in students:
                    ev = sorted(s.evaluations, key=lambda e: (e.review_no, e.id))[-1] if s.evaluations else None
                    if not ev:
                        continue
                    
                    # Calculate individual evaluator totals
                    member1_total = (ev.member1_literature or 0) + (ev.member1_problem or 0) + (ev.member1_presentation or 0) + (ev.member1_qa or 0)
                    member2_total = (ev.member2_literature or 0) + (ev.member2_problem or 0) + (ev.member2_presentation or 0) + (ev.member2_qa or 0)
                    guide_total = (ev.guide_literature or 0) + (ev.guide_problem or 0) + (ev.guide_presentation or 0) + (ev.guide_qa or 0)
                    
                    # Calculate averages
                    avg_literature = int(round(((ev.member1_literature or 0) + (ev.member2_literature or 0) + (ev.guide_literature or 0)) / 3))
                    avg_problem = int(round(((ev.member1_problem or 0) + (ev.member2_problem or 0) + (ev.guide_problem or 0)) / 3))
                    avg_presentation = int(round(((ev.member1_presentation or 0) + (ev.member2_presentation or 0) + (ev.guide_presentation or 0)) / 3))
                    avg_qa = int(round(((ev.member1_qa or 0) + (ev.member2_qa or 0) + (ev.guide_qa or 0)) / 3))
                    avg_total = avg_literature + avg_problem + avg_presentation + avg_qa
                    
                    writer.writerow([
                        s.group_no or "",
                        s.project_title or "",
                        s.seat_no,
                        s.name,
                        ev.review_no,
                        # Member 1 marks
                        ev.member1_literature or 0,
                        ev.member1_problem or 0,
                        ev.member1_presentation or 0,
                        ev.member1_qa or 0,
                        member1_total,
                        # Member 2 marks
                        ev.member2_literature or 0,
                        ev.member2_problem or 0,
                        ev.member2_presentation or 0,
                        ev.member2_qa or 0,
                        member2_total,
                        # Guide marks
                        ev.guide_literature or 0,
                        ev.guide_problem or 0,
                        ev.guide_presentation or 0,
                        ev.guide_qa or 0,
                        guide_total,
                        # Average marks
                        avg_literature,
                        avg_problem,
                        avg_presentation,
                        avg_qa,
                        avg_total,
                    ])
        return send_file(export_path, as_attachment=True, download_name="evaluations_review1.csv", mimetype="text/csv")
    
    @app.route("/summary.pdf")
    def download_summary_pdf():
        # Collect all evaluations grouped by phase and review
        students = Student.query.order_by(Student.name).all()
        
        # Structure: {(phase, review): {'guides': {...}, 'groups': {...}}}
        all_data = {}
        
        # Define all possible phase-review combinations
        phase_review_combos = [(1, 1), (1, 2), (2, 1), (2, 2)]
        
        for phase, review in phase_review_combos:
            guides = {}
            groups = {}
            
            for student in students:
                ev = next((e for e in student.evaluations if e.phase == phase and e.review_no == review), None)
                
                # Add to guides
                if ev and student.project_guide:
                    guide_name = student.project_guide.strip()
                    if guide_name not in guides:
                        guides[guide_name] = []
                    guides[guide_name].append((student, ev))
                
                # Add to groups
                if ev:
                    group_no = student.group_no or "No Group"
                    if group_no not in groups:
                        groups[group_no] = []
                    groups[group_no].append((student, ev))
            
            # Only include if there's data
            if guides or groups:
                all_data[(phase, review)] = {'guides': guides, 'groups': groups}
        
        if not all_data:
            flash("No evaluation data found to generate summary.", "error")
            return redirect(url_for("list_students"))
        
        pdf_buffer = build_comprehensive_pdf(all_data)
        filename = f"CIE_Comprehensive_Report_All_Reviews_{date.today().strftime('%Y%m%d')}.pdf"
        return send_file(pdf_buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")

    return app

if __name__ == "__main__":
    app = create_app()
    app.run(host="127.0.0.1", port=5000, debug=True)
